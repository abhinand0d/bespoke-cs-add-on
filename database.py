import pyodbc
import json
import os
import openpyxl

DBQ = 'D:\BESPOKE TSR\MData.mdb' #DB Location - Change it later
conn = pyodbc.connect(r'Driver={Microsoft Access Driver (*.mdb, *.accdb)};DBQ='+DBQ+';')
LOCATION = os.getenv("LOCATION")


def EXCEL_DATA():
    with open("config.json","r") as f:
        data = json.load(f)
        EXCEL = data["EXCEL"]

    if not os.path.exists(f"{EXCEL}data.xlsx"):
        # Create new excel file
        openpyxl.Workbook().save(f"{EXCEL}data.xlsx")

    try:
        MONTHLY_EXP = openpyxl.load_workbook(f"{EXCEL}data.xlsx")
        data_to_json = {}
        MONTHLY_EXP_SHEET = MONTHLY_EXP.active
        for i in range(1,MONTHLY_EXP_SHEET.max_row+1):
            # check if second value is numeric
            if isinstance(MONTHLY_EXP_SHEET.cell(i,2).value,int) or isinstance(MONTHLY_EXP_SHEET.cell(i,2).value,float):
                data_to_json[MONTHLY_EXP_SHEET.cell(i,1).value] = MONTHLY_EXP_SHEET.cell(i,2).value

        return data_to_json
    except:
        return None

class Bespoke():

    def sale(date):
        sale = {"total_amt": 0,"Location":LOCATION}
        total_amt = 0
        cur = conn.cursor()
        cur.execute(f"SELECT nsaleno,nsaleamt,ntotdiscamt FROM salesmast WHERE dsaledate=#{date}#")
        for i in cur.fetchall():
            sale[i[0]] = {"sale_amount": i[1], "discount_amount": i[2],"purchase_rate":0,"profit":0,"item_list": []}
            try: 
                total_amt += i[1]
            except:
                total_amt += 0
        sale["total_amt"] = total_amt
        if total_amt == 0:
            sale["total_amt"] = 0
            return sale
        else:
            vs = list(sale.keys())
            vs.remove("total_amt")
            vs.remove("Location")
            for iu in vs:
                codes = []
                price = []
                qty = []
                try:
                    cur.execute(f"SELECT citcode,nprice,nqty FROM salestran WHERE nsaleno={iu}")
                    ddp = []
                    purchase_rate = 0
                    profit_rate = 0
                    for xi in cur.fetchall():
                        citcode = xi[0]
                        nprice = xi[1]
                        nqty = xi[2]
                        piprice = nprice/nqty 
                    
                        ITEM_CODE, ITEM_NAME, _, ITEM_SALE_RATE = Bespoke.item(code=citcode)

                        if piprice != ITEM_SALE_RATE: # this will update user if the rate is different 
                            ITEM_PURCHASE_RATE = int(piprice) * 0.7
                            ITEM_SALE_RATE = piprice
                        
                        ITEM_PURCHASE_RATE = round(ITEM_SALE_RATE * 0.7,2)

                        purchase_rate += ITEM_PURCHASE_RATE * nqty
                        PROFIT = (int(ITEM_SALE_RATE) - int(ITEM_PURCHASE_RATE)) * nqty
                        profit_rate += PROFIT

                        ddp.append([citcode,ITEM_NAME,nprice,nqty,ITEM_SALE_RATE,ITEM_PURCHASE_RATE,PROFIT])

                    sale[iu]["item_list"] = ddp
                    sale[iu]["purchase_rate"] = purchase_rate
                    sale[iu]["profit"] = float(profit_rate) - sale[iu]["discount_amount"]


                
                except Exception as e:
                    print(e)
        return sale

    def item(code):
        cur = conn.cursor()
        ITEM_CODE = code
        cur.execute(f"SELECT citname FROM item WHERE citcode='{code}'")
        ITEM_NAME = cur.fetchone()[0]
        cur.execute(f"SELECT nsalerate,nprrate FROM itembatch WHERE citbatcode='{code}'")
        ITEM_SALE_RATE, ITEM_PURCHASE_RATE = cur.fetchone()

        return ITEM_CODE, ITEM_NAME, ITEM_PURCHASE_RATE, ITEM_SALE_RATE
    
    def item_list():
        cur = conn.cursor()
        ITEM_LIST = []
        cur.execute("SELECT citcode FROM item ORDER BY citcode ASC;")
        for i in cur.fetchall():
            ITEM_LIST.append(i[0])

        return ITEM_LIST
